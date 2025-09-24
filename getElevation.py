# Excelのデータから標高を取得し，csvに保存する
import pandas as pd
from pyproj import Transformer
import elevation_service as elevation_service
import os
from openpyxl import load_workbook

# Excelファイルの読み込み
input_excel = "preExpData_fatigue.xlsx"  # 入力ファイル名
sheet_name = "座標取得"  # シート名または番号
# B列（緯度）とC列（経度）だけを読み込む（可変長対応）
df = pd.read_excel(input_excel,sheet_name= sheet_name, usecols="B:C", skiprows=1, header=None)
df.columns = ["latitude", "longitude"]

wb = load_workbook(input_excel)
ws = wb['高さの取得']  # シート名を指定してワークブックを取得

# 緯度・経度の列名（Excelに応じて変更）
lat_col = '緯度'
lon_col = '経度'
count = 1

tif_path = f"{os.path.dirname(os.path.abspath(__file__))}/tool/tif_merged/merged_all.tif"
elevation_service_ins = elevation_service.ElevationService(tif_path)

for lat, lon in zip(df["latitude"], df["longitude"]):
    count+= 1
    elevation = elevation_service_ins.get_elevation(lat, lon)
    # print(f"lat: {lat}, lon: {lon}, elevation: {elevation}")
    ws.cell(row=count, column=2, value=elevation)  

wb.save(input_excel)  # 上書き保存

# # 緯度・経度が存在するか確認
# if lat_col not in df.columns or lon_col not in df.columns:
#     raise ValueError(f"'{lat_col}' または '{lon_col}' が見つかりません")
